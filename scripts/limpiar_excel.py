#!/usr/bin/env python3
"""
limpiar_excel.py — Limpia Excel desde almacenamiento externo (USB/Manjaro).
Copia el archivo a la RAM (/tmp) para evitar saturar el bus USB.
Corrige automáticamente problemas de encoding (UTF-8) y columnas desplazadas.
Uso: python scripts/limpiar_excel.py "data/staging/cammesa_2025/Base de datos Informe Anual 2025.xlsx"
"""

import sys
import os
import shutil
import tempfile
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.chartsheet import Chartsheet


def reparar_encoding_texto(val):
    """Corrige caracteres raros si venían mal codificados en la matriz del Excel."""
    if isinstance(val, str):
        reemplazos = {
            'Ã¡': 'á', 'Ã©': 'é', 'Ã­': 'í', 'Ã³': 'ó', 'Ãº': 'ú',
            'Ã‘': 'Ñ', 'Ã±': 'ñ', 'Ã': 'Á', 'Ã“': 'Ó'
        }
        for roto, sano in reemplazos.items():
            val = val.replace(roto, sano)
        return val.strip()
    return val


def limpiar_excel_usb(archivo_origen):
    if not os.path.exists(archivo_origen):
        print(f"❌ Error: No se encuentra '{archivo_origen}'. Verificá la ruta.")
        sys.exit(1)

    # 1. Crear carpeta de salida en el disco origen
    carpeta_salida = os.path.join(os.path.dirname(archivo_origen), "tablas_limpias")
    os.makedirs(carpeta_salida, exist_ok=True)

    # 2. Copiar el Excel a la memoria RAM (/tmp en Linux) para lectura ultra rápida
    print("⚡ Copiando archivo a memoria RAM (/tmp) para evitar lecturas lentas por USB...")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_path = tmp_file.name

    shutil.copy2(archivo_origen, tmp_path)

    wb = None
    try:
        print("🚀 Procesando archivo desde RAM...")
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

        HOJAS_IGNORAR = {'Contenido', 'INDEX', 'PORTADA'}

        for nombre_hoja in wb.sheetnames:
            if nombre_hoja.strip() in HOJAS_IGNORAR:
                print(f"⏩ Omitiendo portada/índice: '{nombre_hoja}'")
                continue

            print(f"\n--- Procesando hoja: '{nombre_hoja}' ---")

            if isinstance(wb[nombre_hoja], Chartsheet):
                print(f"⏩ Omitiendo hoja de gráfico: '{nombre_hoja}'")
                continue

            hoja = wb[nombre_hoja]

            filas = []
            for fila in hoja.iter_rows(values_only=True):
                # Arreglar caracteres raros celda por celda
                fila_limpia = [reparar_encoding_texto(celda) for celda in fila]
                if any(c is not None and str(c).strip() != "" for c in fila_limpia):
                    filas.append(fila_limpia)

            if not filas:
                print(f"  Hoja '{nombre_hoja}' vacía. Se omite.")
                continue

            df = pd.DataFrame(filas)
            df_clean = df.dropna(how="all", axis=1).reset_index(drop=True)
            df_clean.columns = range(df_clean.shape[1])

            # Detectar bloques de datos
            filas_vacias = df_clean.isnull().all(axis=1)
            indices_corte = np.where(filas_vacias)[0]

            bloques = []
            inicio = 0

            for corte in indices_corte:
                if inicio < corte:
                    sub_df = df_clean.iloc[inicio:corte].dropna(how="all", axis=1).dropna(how="all", axis=0)
                    if not sub_df.empty:
                        bloques.append(sub_df)
                inicio = corte + 1

            if inicio < len(df_clean):
                sub_df = df_clean.iloc[inicio:].dropna(how="all", axis=1).dropna(how="all", axis=0)
                if not sub_df.empty:
                    bloques.append(sub_df)

            hoja_slug = "".join(c if c.isalnum() else "_" for c in nombre_hoja).strip("_")

            for idx, tabla in enumerate(bloques, start=1):
                # Eliminar columnas que sean totalmente vacías en ESTE bloque
                tabla = tabla.dropna(how="all", axis=1).reset_index(drop=True)

                # 1. Buscar el encabezado real (saltar títulos sueltos como "Evolución anual...")
                header_idx = 0
                for i in range(len(tabla)):
                    if tabla.iloc[i].count() > 2:
                        header_idx = i
                        break

                # Setear las columnas correctas y descartar los títulos de arriba
                tabla.columns = tabla.iloc[header_idx]
                tabla = tabla[header_idx + 1:].reset_index(drop=True)

                # 2. "Arrastrar" los nombres de categoría hacia abajo (Forward Fill)
                if not tabla.empty and len(tabla.columns) > 0:
                    tabla.iloc[:, 0] = tabla.iloc[:, 0].ffill()

                base_name = f"{hoja_slug}_tabla_{idx}"
                ruta_csv = os.path.join(carpeta_salida, f"{base_name}.csv")
                ruta_md = os.path.join(carpeta_salida, f"{base_name}.md")

                # Guardar CSV con BOM en UTF-8
                tabla.to_csv(ruta_csv, index=False, encoding="utf-8-sig")

                try:
                    tabla.to_markdown(ruta_md, index=False)
                except ImportError:
                    with open(ruta_md, "w", encoding="utf-8") as f:
                        f.write(tabla.to_csv(sep="|", index=False))

                print(f"  ✔️ Guardado y formateado: {os.path.basename(ruta_csv)}")
    finally:
        if wb is not None:
            wb.close()

        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except PermissionError:
            pass

    print(f"\n✅ ¡Proceso finalizado! Tablas generadas en: {carpeta_salida}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/limpiar_excel.py <archivo.xlsx>")
        sys.exit(1)
    limpiar_excel_usb(sys.argv[1])